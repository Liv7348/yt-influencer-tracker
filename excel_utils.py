"""엑셀 입출력 헬퍼 (천단위 콤마 서식, 템플릿 생성)."""

import io

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from weights import CHANNEL_TYPES

TEMPLATE_HEADERS = [
    "NO",
    "유튜버",
    "채널URL",
    "옵션명",
    "채널유형",
    "평균조회수(비워두면 유튜브는 자동수집)",
    "집행금액(VAT제외, 옵션당 1행만 입력)",
]


def build_excel(sheets: dict, comma_cols: dict, decimal_cols: dict = None) -> bytes:
    """sheets: {시트명: DataFrame}, comma_cols: {시트명: [정수 콤마서식 컬럼]}, decimal_cols: {시트명: [소수1자리 콤마서식 컬럼]}"""
    decimal_cols = decimal_cols or {}
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
            ws = writer.sheets[name[:31]]
            cols = list(df.columns)

            for col_name in comma_cols.get(name, []):
                if col_name not in cols:
                    continue
                col_letter = get_column_letter(cols.index(col_name) + 1)
                for row in range(2, len(df) + 2):
                    ws[f"{col_letter}{row}"].number_format = "#,##0"

            for col_name in decimal_cols.get(name, []):
                if col_name not in cols:
                    continue
                col_letter = get_column_letter(cols.index(col_name) + 1)
                for row in range(2, len(df) + 2):
                    ws[f"{col_letter}{row}"].number_format = "#,##0.0"

            for i, col_name in enumerate(cols, start=1):
                values = df[col_name].astype(str).tolist()
                max_len = max([len(str(col_name))] + [len(v) for v in values]) if values else len(str(col_name))
                ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)

    buffer.seek(0)
    return buffer.getvalue()


def build_template_excel() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "가중단가_입력"
    ws.append(TEMPLATE_HEADERS)

    example_rows = [
        [1, "예시크리에이터", "https://www.youtube.com/@example", "롱폼", "유튜브 롱폼", "", 30000000],
        [2, "예시크리에이터", "https://www.youtube.com/@example", "롱폼+숏츠 미러링", "유튜브 롱폼", "", 35000000],
        ["", "", "", "", "유튜브 숏츠", "", ""],
        [3, "예시크리에이터2", "", "롱폼+릴스 미러링", "인스타 릴스", 80000, ""],
    ]
    for row in example_rows:
        ws.append(row)

    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(CHANNEL_TYPES)}"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add("E2:E500")

    for i, header in enumerate(TEMPLATE_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(18, len(header) + 2)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def comma(n) -> str:
    if n is None or (isinstance(n, float) and pd.isna(n)):
        return "N/A"
    return f"{int(round(n)):,}"
